import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import json

# Obtém o nome do usuário logado
username = os.getlogin()

# URL da API (substituída por um valor genérico)
url = "https://seudominio.com.br:porta/api/framework/v1/consultaSQLServer/RealizaConsulta/TIPES0001/0/P"

# Credenciais de autenticação básica (substituídas por valores genéricos)
Login = "seu_usuario"
password = "sua_senha"

# Fazendo a requisição GET com autenticação básica
response = requests.get(url, auth=HTTPBasicAuth(Login, password))

# Verificando se a requisição foi bem-sucedida
if response.status_code == 200:
    dados = response.json()
    df = pd.DataFrame(dados)

    # Função para selecionar o melhor registro por CPF
    def selecionar_melhor_registro(grupo):
        if len(grupo) == 1:
            return grupo
        if 'Ativo' in grupo['SITUACAO'].values and 'Demitido' in grupo['SITUACAO'].values:
            grupo = grupo[grupo['SITUACAO'] == 'Ativo']
        if len(grupo) > 1:
            grupo = grupo.sort_values(by=['EMAIL'], ascending=False)
        return grupo.head(1)

    df = df.groupby('CPF', group_keys=False).apply(selecionar_melhor_registro).reset_index(drop=True)
    df.loc[df["CPF"] == "99999999999", "NOME"] = "Suporte"

    df = df.sort_values(by="NOME", ascending=True).reset_index(drop=True)

    # Caminho de destino
    caminho_downloads = f"C:\\Users\\{username}\\...\\03 Base_Inventario"
    nome_arquivo_excel = "Colaboradores_Administrativos.xlsx"
    caminho_completo_excel = os.path.join(caminho_downloads, nome_arquivo_excel)

    # Remover o arquivo se já existir
    if os.path.exists(caminho_completo_excel):
        os.remove(caminho_completo_excel)

    # Salvar como Excel
    df.to_excel(caminho_completo_excel, index=False)

    # Adicionando formatação de tabela no Excel
    workbook = load_workbook(caminho_completo_excel)
    sheet = workbook.active
    max_row, max_col = sheet.max_row, sheet.max_column
    table_range = f"A1:{chr(64 + max_col)}{max_row}"
    tab = Table(displayName="Dcolab", ref=table_range)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
    sheet.add_table(tab)
    workbook.save(caminho_completo_excel)

    print(f"Dados salvos com sucesso no arquivo: {caminho_completo_excel}")

    # Exportar parte dos dados para JSON
    nome_arquivo_json = "Dcolab.json"
    caminho_completo_json = os.path.join(caminho_downloads, nome_arquivo_json)
    df_json = df[["CPF", "NOME", "EMAIL", "DESCRICAO_SECAO", "FUNCAO", "SITUACAO", "NM_EMPRESA", "DT_ADMISSAO", "DT_DEMISSAO", "CODSECAO","RE","CHAPA"]]
    df_json.to_json(caminho_completo_json, orient="records", force_ascii=False, indent=4)

    print(f"Arquivo JSON gerado com sucesso: {caminho_completo_json}")

else:
    print(f"Erro na requisição: {response.status_code}")
    print("Resposta:", response.text)
